import base64
from datetime import timezone
from pickle import GET
from urllib import request
from django.core.mail import send_mail
from django.conf import settings
from django.urls import reverse
from django.utils.html import strip_tags
import openpyxl
from django.shortcuts import render,redirect
from django.contrib.auth import logout
from sdnts.models import CategoriaInsumo, DetalleVenta, Entrada, Envio, Produccion, Proveedor, Salida, Usuario, Producto, Carrito, CarritoItem, Venta, Domiciliario, Cliente, Pago, Notificacion
from django.contrib.auth import views as auth_views
from django.urls import reverse, reverse_lazy
import json
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .forms import CargarDatosForm, PerfilAdminForm, CambiarContrasenaForm, RegistroUsuarioForm, UsuarioForm # Asume que tienes este formulario creado
from django.contrib.auth import login as auth_login
from django.contrib.auth.decorators import login_required
from sdnts.models import SaborMasa, Glaseado, Topping, CombinacionProducto, Usuario
import io
import matplotlib.pyplot as plt
from django.http import HttpResponse
from weasyprint import HTML
from django.template.loader import render_to_string
from reportlab.pdfgen import canvas
from django.db.models import Sum

from decimal import Decimal

from django.utils.timezone import now
from datetime import timedelta
from django.contrib.auth import update_session_auth_hash
from django.utils import timezone
from datetime import timedelta


from django.contrib.auth import get_user_model

# Context processors
def nav_index(request):
    return render(request, 'includes/nav_index.html')

def nav_admin(request):
    return render(request, 'includes/nav_admin.html')

def nav_user(request):
    return render(request, 'includes/nav_user.html')


# --- FUNCION NORMALIZAR PARA MAPEOS DE NOMBRES ---
def normalizar(nombre):
    if not nombre:
        return ''
    import unicodedata
    nombre = nombre.lower().strip()
    nombre = ''.join(
        c for c in unicodedata.normalize('NFD', nombre)
        if unicodedata.category(c) != 'Mn'
    )
    nombre = nombre.replace('-', ' ').replace('_', ' ')
    nombre = nombre.replace('m&m', 'mm')
    return nombre

# Create your views here.
def index(request):
    return render(request, 'index/index.html')

def catalogo(request):
    return render(request, 'index/catalogo.html')

def contactanos(request):
    context = {
        'direccion': 'Cra 13 # 6510, Bogotá, Colombia',
        'horario': 'Lunes a Sábado 9:00 a.m - 6:00 p.m',
        'telefono': '+57 3026982043'
    }
    return render(request, 'index/contactanos.html', context)
def nosotros(request):
    
    return render(request, 'index/nosotros.html')


def login(request):
    if request.method == 'POST':
        email = request.POST['email']
        password = request.POST['password']
        
        usuario = Usuario.objects.filter(email=email).first()

        if usuario and usuario.check_password(password):
            # Guardamos el ID del usuario en sesión
            request.session['cod_usuario'] = usuario.cod_usuario
            
            # También puedes iniciar sesión con Django si quieres
            auth_login(request, usuario)
            
            rol = usuario.rol.upper() if usuario.rol else ''

            # Redirección por rol
            if rol == 'ADMIN':
                return redirect('dashboard_admin')
            elif rol == 'CLIENTE':
                return redirect('vistacliente')  
            elif rol == 'DOMI':
                return redirect('mis_domicilios')
            else:
                return redirect('error_view')
        else:
            return render(request, 'auth/login.html', {'error': 'Credenciales inválidas'})
    
    return render(request, 'auth/login.html')


def registro(request):
    if request.method == 'POST':
        form = RegistroUsuarioForm(request.POST)  # Este formulario NO tiene campo 'rol'
        if form.is_valid():
            usuario = form.save(commit=False)
            usuario.rol = 'CLIENTE'  # Asigna el rol por defecto
            usuario.save()
            enviar_correo_bienvenida(usuario)
            return redirect('login')
    else:
        form = RegistroUsuarioForm()
    return render(request, 'auth/registro.html', {'form': form})


@login_required
def agregar_usuario(request):
    if request.method == 'POST':
        form = UsuarioForm(request.POST)  # Este formulario SÍ tiene campo 'rol'
        if form.is_valid():
            usuario = form.save()
            enviar_correo_bienvenida_admin(usuario)
            messages.success(request, "Usuario agregado exitosamente.")
            return redirect('dashboard_admin')
        else:
         messages.error(request, 'Corrige los errores en el formulario.')
    else: 
        form = UsuarioForm()
    return render(request, 'usuario/agregar_usuario.html', {'form': form})


class CustomPasswordResetView(auth_views.PasswordResetView):
    template_name = 'auth/password_reset.html'
    extra_context = {'etapa': 'formulario'}

class CustomPasswordResetDoneView(auth_views.PasswordResetDoneView):
    template_name = 'auth/password_reset.html'
    extra_context = {'etapa': 'correo_enviado'}

class CustomPasswordResetConfirmView(auth_views.PasswordResetConfirmView):
    template_name = 'auth/password_reset.html'
    extra_context = {'etapa': 'nueva_contraseña'}

class CustomPasswordResetCompleteView(auth_views.PasswordResetCompleteView):
    template_name = 'auth/password_reset.html'
    extra_context = {'etapa': 'completado'}
        
    
def enviar_correo_bienvenida(usuario):
    asunto = '¡Bienvenido a StefasDonuts!'
    mensaje_html = f'''
    <h2>Hola {usuario.nom_usua}!</h2>
    <p>Te damos la bienvenida a <strong>StefasDonutsm</strong>.</p>
    <p>Tus datos son:</p>
    <ul>
        <li><strong>Usuario:</strong> {usuario.nom_usua}</li>
        <li><strong>Email:</strong> {usuario.email}</li>
    </ul>
    <p>Puedes iniciar sesión aquí: 
        <a href="http://127.0.0.1:8000/{reverse('login')}">Iniciar sesión</a>
    </p>
    '''
    mensaje_texto = strip_tags(mensaje_html)  # Por si el cliente de correo no soporta HTML

    send_mail(
        asunto,
        mensaje_texto,
        settings.DEFAULT_FROM_EMAIL,
        [usuario.email],
        html_message=mensaje_html
    )
    
def enviar_correo_bienvenida_admin(usuario):
    asunto = '¡Bienvenido a StefasDonuts!'
    mensaje_html = f'''
    <h2>Hola {usuario.nom_usua}!</h2>
    <p>Te damos la bienvenida a <strong>StefasDonuts</strong>.</p>
    <p>Tus datos son:</p>
    <ul>
        <li><strong>Usuario:</strong> {usuario.nom_usua}</li>
        <li><strong>Email:</strong> {usuario.email}</li>
        <li><strong>Rol asignado:</strong> {usuario.rol}</li>
    </ul>
    <p>Puedes iniciar sesión aquí: 
        <a href="http://127.0.0.1:8000{reverse('login')}">Iniciar sesión</a>
    </p>
    '''
    mensaje_texto = strip_tags(mensaje_html)
    send_mail(
        asunto,
        mensaje_texto,
        settings.DEFAULT_FROM_EMAIL,
        [usuario.email],
        html_message=mensaje_html
    )
    
    
@login_required   
def logout_view(request):
    logout(request)
    return redirect('login') 
@login_required    
def vistacliente(request):
    return render(request, 'cliente/vistacliente.html') 
@login_required
def catalogocliente(request):
    masas = SaborMasa.objects.all()
    coberturas = Glaseado.objects.all()
    toppings = Topping.objects.all()

    # Obtener productos por tamaño
    producto_s = Producto.objects.filter(tamano='S', activo=True).first()
    producto_m = Producto.objects.filter(tamano='M', activo=True).first()
    producto_l = Producto.objects.filter(tamano='L', activo=True).first()
    producto_xl = Producto.objects.filter(tamano='XL', activo=True).first()

    return render(request, 'cliente/catalogocliente.html', {
        'masas': masas,
        'coberturas': coberturas,
        'toppings': toppings,
        'producto_s': producto_s,
        'producto_m': producto_m,
        'producto_l': producto_l,
        'producto_xl': producto_xl,
        # ...otros contextos...
    })
    return render(request, 'cliente/catalogocliente.html')

@login_required
def contactanoscliente(request):
    context = {
        'direccion': 'Cra 13 # 6510, Bogotá, Colombia',
        'horario': 'Lunes a Sábado 9:00 a.m - 6:00 p.m',
        'telefono': '+57 3026982043'
    }
    return render(request, 'cliente/contactanoscliente.html', context)

@login_required
def pago(request):
    """Vista para la página de pago/factura"""
    return render(request, 'cliente/pago.html')

def nosotroscliente(request):
    return render(request, 'cliente/nosotroscliente.html')

from django.contrib import messages
from django.shortcuts import redirect, render
from django.contrib.auth.decorators import login_required
from .models import Cliente, Venta

@login_required
def perfilcliente(request):
    try:
        usuario = request.user
        cliente = Cliente.objects.get(cod_usua=usuario)

        ventas = Venta.objects.filter(
            cod_cliente=cliente
        ).prefetch_related(
            'detalles',
            'detalles__combinaciones',
            'detalles__combinaciones__cod_producto',
            'detalles__combinaciones__cod_sabor_masa_1',
            'detalles__combinaciones__cod_glaseado_1',
            'detalles__combinaciones__cod_topping_1',
        ).order_by('-fecha_hora')

        context = {
            'usuario': usuario,
            'cliente': cliente,
            'ventas': ventas,
        }

        return render(request, 'cliente/perfilcliente.html', context)
    except Exception as e:
        print(f"Error en perfilcliente: {str(e)}")
        messages.error(request, 'Error al cargar el perfil')
        return redirect('vistacliente')

@login_required
def editar_perfil(request):
    if request.method == 'POST':
        # Procesar el formulario
        request.user.nom_usua = request.POST.get('nombre')
        request.user.apell_usua = request.POST.get('apellido')
        request.user.email = request.POST.get('email')
        request.user.tele_usua = request.POST.get('telefono')
        request.user.cliente.direc_cliente = request.POST.get('direccion')
        
        try:
            request.user.save()
            request.user.cliente.save()
            messages.success(request, 'Perfil actualizado correctamente')
            return redirect('perfilcliente')
        except Exception as e:
            messages.error(request, f'Error al actualizar: {str(e)}')
    
    # Para la Opción 2:
    editando = request.GET.get('editar') == '1'
    
    return render(request, 'cliente/editar_perfil.html' if not editando else 'perfilcliente.html', {
        'user': request.user,
        'editando': editando  # Solo para Opción 2
    })
    
@login_required
def agregar_usuario(request):
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            usuario = form.save(commit=False)
            usuario.set_password(form.cleaned_data['password1'])
            usuario.save()
            messages.success(request, 'Usuario agregado exitosamente.')
            return redirect('dashboard_admin')
        else:
            messages.error(request, 'Corrige los errores en el formulario.')
    else:
        form = UsuarioForm()
    return render(request, 'usuarios/agregar_usuario.html', {'form': form})

@login_required
def agregar_al_carrito(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        producto_id = data.get('producto_id')
        cantidad = int(data.get('cantidad', 1))
        masa = data.get('masa', {}).get('nombre')
        cobertura = data.get('cobertura', {}).get('nombre')
        topping = data.get('topping', {}).get('nombre')
        
        try:
            producto = get_object_or_404(Producto, cod_producto=producto_id)
            
            # Obtener o crear el carrito para el usuario
            carrito, _ = Carrito.objects.get_or_create(
                usuario=request.user,
                completado=False
            )
            
            # Buscar si existe un item con las mismas características
            item = CarritoItem.objects.filter(
                carrito=carrito,
                producto=producto,
                masa_seleccionada=masa,
                cobertura_seleccionada=cobertura,
                topping_seleccionado=topping
            ).first()
            
            if item:
                # Si existe, actualizar cantidad
                item.cantidad += cantidad
                item.save()
            else:
                # Si no existe, crear nuevo item
                item = CarritoItem.objects.create(
                    carrito=carrito,
                    producto=producto,
                    cantidad=cantidad,
                    masa_seleccionada=masa,
                    cobertura_seleccionada=cobertura,
                    topping_seleccionado=topping
                )
            
            return JsonResponse({
                'success': True,
                'mensaje': 'Producto agregado al carrito',
                'total_items': carrito.cantidad_items,
                'total_carrito': float(carrito.total)
            })
            
        except Exception as e:
            print(f"Error: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

@login_required
def eliminar_del_carrito(request, item_id):
    try:
        item = get_object_or_404(CarritoItem, id=item_id, carrito__usuario=request.user)
        carrito = item.carrito
        item.delete()
        
        return JsonResponse({
            'success': True,
            'mensaje': 'Item eliminado del carrito',
            'total_items': carrito.cantidad_items,
            'total_carrito': float(carrito.total)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
def ver_carrito(request):
    try:
        carrito = Carrito.objects.filter(
            usuario=request.user,
            completado=False
        ).prefetch_related('items', 'items__producto').first()
        
        items = []
        if carrito:
            items = CarritoItem.objects.filter(carrito=carrito).select_related('producto')
        
        context = {
            'carrito': carrito,
            'items': items,
        }
        
        return render(request, 'includes/carrito.html', context)
        
    except Exception as e:
        messages.error(request, f'Error al cargar el carrito: {str(e)}')
        return redirect('catalogocliente')

@login_required
def actualizar_carrito(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            item_id = data.get('item_id')
            cantidad = int(data.get('cantidad', 1))
            
            item = get_object_or_404(CarritoItem, 
                                   id=item_id, 
                                   carrito__usuario=request.user)
            
            if cantidad > 0:
                item.cantidad = cantidad
                item.save()
            else:
                item.delete()
            
            carrito = item.carrito
            
            return JsonResponse({
                'success': True,
                'subtotal': float(item.subtotal()),
                'total_carrito': float(carrito.total),
                'total_items': carrito.cantidad_items
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
            
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.http import JsonResponse
import traceback


@login_required
@csrf_exempt
def procesar_compra(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            carrito = data.get('carrito', [])
            direccion = data.get('direccion', '')
            fecha_entrega = data.get('fecha_entrega')
            metodo_pago = data.get('metodo_pago', 'NEQUI')
            transaccion_id = data.get('transaccion_id', '')

            user = request.user
            cliente = Cliente.objects.get(cod_usua=user)

            subtotal = sum(item['precio'] * item['quantity'] for item in carrito)
            iva = subtotal * 0.19
            total = subtotal + iva

            venta = Venta.objects.create(
                cod_cliente=cliente,
                subtotal=subtotal,
                iva=iva,
                total=total,
                direccion_entrega=direccion,
                fecha_hora=timezone.now()
            )

            # Mapeo para buscar ingredientes correctamente
            def normalizar(texto):
                return texto.lower().strip().replace(" ", "").replace(".", "").replace("-", "")

            MASA_MAP = {
                normalizar("Vainilla"): "Vainilla",
                normalizar("Chocolate"): "Chocolate",
                normalizar("Red Velvet"): "Red Velvet",
            }
            COBERTURA_MAP = {
                normalizar("Chocolate Blanco"): "Choc. Blanco",
                normalizar("Chocolate Oscuro"): "Choc. Oscuro",
                normalizar("Arequipe"): "Arequipe",
            }
            TOPPING_MAP = {
                normalizar("Chispas"): "Chispas",
                normalizar("Oreo"): "Oreo",
                normalizar("M&M"): "M&M",
                normalizar("mm"): "M&M",
                normalizar("Chips"): "Chips",
                normalizar("Ninguno"): None,
            }

            for item in carrito:
                cod_prod = item['cod_producto']
                producto = Producto.objects.get(cod_producto=cod_prod)

                # Normalizar la fecha_entrega si viene como string
                fecha_entrega_obj = timezone.make_aware(datetime.strptime(fecha_entrega, "%Y-%m-%d"))

                # Crear el detalle de venta
                detalle = DetalleVenta.objects.create(
                    cod_venta=venta,
                    cod_producto=producto,
                    cantidad=item['quantity'],
                    precio_unitario=item['precio'],
                    fecha_entrega=fecha_entrega_obj
                )

                # Obtener ingredientes
                masa_nombre = item['masa']['nombre']
                masa = SaborMasa.objects.filter(nombre__iexact=MASA_MAP.get(normalizar(masa_nombre), masa_nombre)).first()

                cobertura_nombre = item['cobertura']['nombre']
                cobertura = Glaseado.objects.filter(nombre__iexact=COBERTURA_MAP.get(normalizar(cobertura_nombre), cobertura_nombre)).first()

                topping = None
                topping_nombre = item['topping']['nombre']
                if topping_nombre.lower() != 'ninguno':
                    topping = Topping.objects.filter(nombre__iexact=TOPPING_MAP.get(normalizar(topping_nombre), topping_nombre)).first()

                # Crear una combinación personalizada por cada unidad
                for _ in range(item['quantity']):
                 CombinacionProducto.objects.create(
                   cod_venta=venta,  # ← esto es lo que te faltaba
                   cod_detalle=detalle,
                   cod_producto=producto,
                   cod_sabor_masa_1=masa,
                   cod_glaseado_1=cobertura,
                   cod_topping_1=topping
                )


            # Registrar el pago
            Pago.objects.create(
                cod_venta=venta,
                metodo_pago=metodo_pago,
                monto_total=total,
                fecha_hora_pago=timezone.now(),
                estado_pago='PENDIENTE',
                transaccion_id=transaccion_id
            )

            # Notificación para los administradores
            admins = Usuario.objects.filter(rol='ADMIN')
            mensaje = f"Nuevo pedido realizado por {user.nom_usua} {user.apell_usua}."
            for admin in admins:
                Notificacion.objects.create(usuario=admin, mensaje=mensaje)

            return JsonResponse({
                'success': True,
                'venta': {
                    'fecha': venta.fecha_hora.strftime('%d/%m/%Y %H:%M:%S'),
                    'direccion': venta.direccion_entrega,
                    'subtotal': float(venta.subtotal),
                    'iva': float(venta.iva),
                    'total': float(venta.total),
                }
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


from django.db.models import Q  # 👈 Importar para excluir por estado
from django.utils.timezone import now
from datetime import timedelta

@login_required
def mis_domicilios(request):
    try:
        domiciliario = request.user.domiciliario
    except Domiciliario.DoesNotExist:
        messages.error(request, "No tienes un perfil de domiciliario asignado.")
        return render(request, 'domiciliario/mis_domicilios.html', {'envios': [], 'nuevos_envios': []})

    # Mostrar todos excepto ENTREGADO
    envios = Envio.objects.filter(cod_domi=domiciliario).exclude(estado="ENTREGADO")

    recientes = now() - timedelta(days=1)
    nuevos_envios = Envio.objects.filter(
        cod_domi=domiciliario,
        fecha_asignacion__gte=recientes
    ).exclude(estado="ENTREGADO")

    context = {
        'envios': envios,
        'nuevos_envios': nuevos_envios,
    }

    return render(request, 'domiciliario/mis_domicilios.html', context)

@login_required
def historial_envios(request):
    usuario = request.user  # Asegúrate de usar autenticación
    fecha_desde = request.GET.get('desde')
    fecha_hasta = request.GET.get('hasta')
    
    envios = Envio.objects.filter(cod_domi= request.user.domiciliario).order_by('fecha_entrega')

    nuevos_envios = Envio.objects.filter(
    cod_domi=usuario.domiciliario,
    fecha_asignacion__gte=timezone.now() - timedelta(days=1)
    ).order_by('-fecha_asignacion')


    return render(request, 'domiciliario/historial_envios.html', {
        'envios': envios,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'nuevos_envios': nuevos_envios,  # ✅ se pasa al template
    })
  
from django.shortcuts import get_object_or_404
from django.contrib import messages
from sdnts.forms import EnvioDomiciliarioForm


def editar_envio_domiciliario(request, cod_envio):
    envio = get_object_or_404(Envio, cod_envio=cod_envio)

    if request.user != envio.cod_domi.cod_usua:
        messages.error(request, "No tienes permiso para editar este envío.")
        return redirect('mis_domicilios')

    estado_anterior = envio.estado  # ← capturar antes de guardar

    if request.method == 'POST':
        form = EnvioDomiciliarioForm(request.POST, instance=envio)
        if form.is_valid():
            envio = form.save(commit=False)

            if envio.estado == "ENTREGADO":
                envio.cod_venta.estado = "ENTREGADO"
                envio.cod_venta.save()
                envio.save()
                messages.success(request, "El envío fue entregado y archivado.")
                return redirect('historial_envios')

            elif envio.estado == "CANCELADO":
                envio.cod_venta.estado = "CANCELADO"
                envio.cod_venta.save()
                envio.save()
                messages.success(request, "El envío fue cancelado.")
                return redirect('mis_domicilios')

            # ⚠️ Solo guarda cambios y no archiva si NO se cambió a entregado
            envio.save()

            if estado_anterior != envio.estado:
                messages.success(request, "Estado actualizado correctamente.")
            else:
                messages.success(request, "Observaciones o fechas actualizadas.")

            return redirect('mis_domicilios')
    else:
        form = EnvioDomiciliarioForm(instance=envio)

    return render(request, 'domiciliario/editar_envio_domiciliario.html', {
        'form': form,
        'envio': envio,
    })

@login_required
def detalle_venta_domiciliario(request, cod_venta):
    venta = get_object_or_404(Venta, cod_venta=cod_venta)
    detalles = DetalleVenta.objects.filter(cod_venta=venta)

    return render(request, 'admin/ventas/detalle_ventas.html', {
        'venta': venta,
        'detalles': detalles,
        'origen': 'mis_domicilios'  # Este es el nombre de la URL de retorno
    })

@login_required
def perfildomi(request):
    user = request.user
    return render(request, 'domiciliario/perfildomi.html', {'user': user})


User = get_user_model()

@login_required
def editar_perfildomi(request):
    user = request.user

    if request.method == 'POST':
        nom_usua = request.POST.get('nom_usua', user.nom_usua)
        apell_usua = request.POST.get('apell_usua', user.apell_usua)
        tele_usua = request.POST.get('tele_usua', user.tele_usua)
        email = request.POST.get('email', user.email)
        password = request.POST.get('password')
        password2 = request.POST.get('confirm_password')

        # Validación: verificar si el email ya existe y no es el del usuario actual
        if User.objects.filter(email=email).exclude(pk=user.pk).exists():
            messages.error(request, 'El correo electrónico ya está en uso.')
            return redirect('perfildomi')

        # Validación: confirmar contraseña si se escribe
        if password:
            if password != password2:
                messages.error(request, 'Las contraseñas no coinciden.')
                return redirect('perfildomi')
            user.set_password(password)
            update_session_auth_hash(request, user)  # para que no cierre la sesión

        # Guardar datos
        user.nom_usua = nom_usua
        user.apell_usua = apell_usua
        user.tele_usua = tele_usua
        user.email = email
        user.save()

        messages.success(request, 'Perfil actualizado correctamente.')
        return redirect('perfildomi')

    return render(request, 'domiciliario/editar_perfildomi.html', {'user': user})


#administrador
@login_required
def editar_usuario(request, cod_usuario):
    usuario = get_object_or_404(Usuario, cod_usuario=cod_usuario)
    rol_anterior = usuario.rol  # Guardamos el rol actual antes del cambio

    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=usuario)

        if form.is_valid():
            usuario_actualizado = form.save(commit=False)  # Aún no guardamos en BD
            nuevo_rol = usuario_actualizado.rol

            # Si el rol cambió, eliminamos el modelo anterior relacionado
            if rol_anterior != nuevo_rol:
                # Eliminamos el modelo del rol anterior
                try:
                    if rol_anterior == 'CLIENTE' and hasattr(usuario, 'cliente'):
                        usuario.cliente.delete()
                    elif rol_anterior == 'DOMI' and hasattr(usuario, 'domiciliario'):
                        usuario.domiciliario.delete()
                    elif rol_anterior == 'ADMIN' and hasattr(usuario, 'administrador'):
                        usuario.administrador.delete()
                except Exception as e:
                    print("Error al eliminar modelo anterior:", e)

            # Guardamos el usuario actualizado (ya con nuevo rol)
            usuario_actualizado.save()

            # Si el rol cambió, creamos el modelo nuevo
            if rol_anterior != nuevo_rol:
                try:
                    if nuevo_rol == 'CLIENTE':
                        Cliente.objects.create(cod_usua=usuario_actualizado, direc_cliente='Sin dirección')
                    elif nuevo_rol == 'DOMI':
                        Domiciliario.objects.create(cod_usua=usuario_actualizado)
                    elif nuevo_rol == 'ADMIN':
                        Administrador.objects.create(cod_usua=usuario_actualizado, estado_admin='ACTIVO')
                except Exception as e:
                    print("Error al crear modelo nuevo:", e)

            messages.success(request, "Usuario actualizado correctamente.")
            return redirect('dashboard_admin')
    else:
        form = UsuarioForm(instance=usuario)

    return render(request, 'usuario/editar_usuario.html', {'form': form, 'usuario': usuario})


@login_required
def agregar_usuario(request):
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario agregado exitosamente.")
            return redirect('dashboard_admin')  # Redirige al dashboard admin
    else:
        form = UsuarioForm()

    return render(request, 'usuario/agregar_usuario.html', {'form': form})

@login_required
def dashboard_admin(request):
    # Conteos principales
    total_usuarios = Usuario.objects.count()
    total_clientes = Usuario.objects.filter(rol='CLIENTE').count()
    total_domis = Usuario.objects.filter(rol='DOMI').count()
    total_ventas = Venta.objects.count()
    monto_total_ventas = Venta.objects.aggregate(total=Sum('total'))['total'] or 0

    ventas_recientes = Venta.objects.select_related('cod_cliente').order_by('-fecha_hora')[:5]
    produccion_reciente = Produccion.objects.select_related('cod_venta').order_by('-fecha_inicio')[:3]

    rol = request.GET.get('rol')
    page = request.GET.get('page', 1)
    usuarios_qs = Usuario.objects.filter(rol=rol) if rol else Usuario.objects.all()
    paginator = Paginator(usuarios_qs.order_by('cod_usuario'), 10)  # 10 usuarios por página
    usuarios = paginator.get_page(page)
    rol_actual = rol  # para mantener el filtro en la paginación

    # Gráfica: Ventas por mes (últimos 6 meses, por monto)
    from collections import OrderedDict
    import calendar
    from django.utils import timezone

    hoy = timezone.now()
    meses = []
    for i in range(5, -1, -1):
        y = hoy.year
        m = hoy.month - i
        if m <= 0:
            y -= 1
            m += 12
        meses.append((y, m))
    ventas_mensuales = OrderedDict()
    for y, m in meses:
        label = f"{calendar.month_abbr[m]} {y}"
        ventas_mensuales[label] = 0
    ventas_qs = Venta.objects.filter(fecha_hora__gte=hoy.replace(day=1) - timezone.timedelta(days=180))
    for v in ventas_qs:
        label = f"{calendar.month_abbr[v.fecha_hora.month]} {v.fecha_hora.year}"
        if label in ventas_mensuales:
            ventas_mensuales[label] += float(v.total)
    ventas_mensuales_data = {
        "labels": list(ventas_mensuales.keys()),
        "data": list(ventas_mensuales.values())
    }

    # Gráfica: Ventas por producto (cantidad vendida)
    ventas_por_producto = (
        DetalleVenta.objects.values('cod_producto__nomb_pro')
        .annotate(total=Sum('cantidad'))
        .order_by('-total')
    )
    ventas_por_producto_data = {
        "labels": [v['cod_producto__nomb_pro'] for v in ventas_por_producto],
        "data": [v['total'] for v in ventas_por_producto]
    }

    # Gráfica: Clientes con más compras (top 7)
    from django.db.models import Count
    clientes_top_qs = (
        Venta.objects.values('cod_cliente__cod_usua__nom_usua', 'cod_cliente__cod_usua__apell_usua')
        .annotate(compras=Count('cod_venta'))
        .order_by('-compras')[:7]
    )
    clientes_top_labels = [
        f"{c['cod_cliente__cod_usua__nom_usua']} {c['cod_cliente__cod_usua__apell_usua']}"
        for c in clientes_top_qs
    ]
    clientes_top_data = {
        "labels": clientes_top_labels,
        "data": [c['compras'] for c in clientes_top_qs]
    }

    # Agrega esto para obtener las nuevas ventas (por ejemplo, ventas pendientes)
    nuevas_ventas = Venta.objects.filter(estado='PENDIENTE').order_by('-fecha_hora')[:5]

    # Alerta por insumos próximos a vencer
    hoy_date = timezone.now().date()
    dias_alerta = 7
    insumos_por_vencer = Insumo.objects.filter(fecha_vencimiento__lte=hoy_date + timedelta(days=dias_alerta))

    # Enviar correo si hay insumos por vencer (solo una vez por sesión)
    if insumos_por_vencer.exists() and not request.session.get('correo_insumos_vencer_enviado'):
        lista = "\n".join([f"- {i.nomb_insumo} (vence el {i.fecha_vencimiento})" for i in insumos_por_vencer])
        try:
            send_mail(
                '⚠ Alerta de Insumos por Vencer - KAEF',
                f'Los siguientes insumos están por vencer en los próximos {dias_alerta} días:\n\n{lista}',
                settings.DEFAULT_FROM_EMAIL,  # Usa el correo configurado en settings.py
                ['tucorreo@ejemplo.com'],     # Cambia esto por tu correo real
                fail_silently=False           # Para ver errores si los hay
            )
            request.session['correo_insumos_vencer_enviado'] = True
        except Exception as e:
            messages.warning(request, f'No se pudo enviar el correo de alerta: {e}')

    return render(request, 'admin/dashboard_admin.html', {
        'nuevas_ventas': nuevas_ventas,
        'insumos_por_vencer': insumos_por_vencer,
        'total_usuarios': total_usuarios,
        'total_clientes': total_clientes,
        'total_domis': total_domis,
        'total_ventas': total_ventas,
        'total_ventas_monto': monto_total_ventas,
        'usuarios': usuarios,
        'rol_actual': rol_actual,
        'ventas_mensuales': ventas_mensuales_data,
        'ventas_por_producto': ventas_por_producto_data,
        'clientes_top': clientes_top_data,
    })
    
    
@login_required
def perfil_admin(request):
    usuario = request.user
    return render(request, 'admin/perfil_admin.html', {'usuario': usuario})

@login_required
def editarperfil_admin(request):
    usuario = request.user
    perfil_form = PerfilAdminForm(instance=usuario)
    password_form = CambiarContrasenaForm(user=usuario)

    if request.method == 'POST':
        if 'guardar_perfil' in request.POST:
            perfil_form = PerfilAdminForm(request.POST, instance=usuario)
            if perfil_form.is_valid():
                perfil_form.save()
                messages.success(request, 'Perfil actualizado correctamente.')
                return redirect('perfil_admin')
        
        elif 'cambiar_contrasena' in request.POST:
            password_form = CambiarContrasenaForm(user=usuario, data=request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)  # para que no cierre la sesión
                messages.success(request, 'Contraseña cambiada correctamente.')
                return redirect('perfil_admin')
            else:
                messages.error(request, 'Corrige los errores en el formulario de contraseña.')

    return render(request, 'admin/editarperfil_admin.html', {
        'perfil_form': perfil_form,
        'password_form': password_form
    })
    
# Vista de Ventas
from django.core.paginator import Paginator
from django.db.models import Sum

@login_required
def ventas_admin(request):
    # Mostrar todos los registros de ventas ordenados por cod_venta ascendente
    ventas_list = Venta.objects.select_related('produccion').all().order_by('cod_venta')

    # Si tienes más de 10 registros y solo ves 10, es por la paginación:
    # paginator = Paginator(ventas_list, 10)  # 10 ventas por página
    # page_number = request.GET.get('page', 1)
    # ventas = paginator.get_page(page_number)

    # Para mostrar TODOS los registros en una sola página, elimina la paginación:
    ventas = ventas_list  # Sin paginación

    # Estadísticas
    ventas_pendientes = Venta.objects.filter(estado='PENDIENTE').count()
    ventas_en_camino = Venta.objects.filter(estado='EN_CAMINO').count()
    ventas_entregadas = Venta.objects.filter(estado='ENTREGADO').count()

    total_ventas = ventas_list.aggregate(Sum('total'))['total__sum'] or 0

    context = {
        'ventas': ventas,
        'total_ventas': total_ventas,
        'ventasPendientes': ventas_pendientes,
        'ventasEnCamino': ventas_en_camino,
        'ventasEntregadas': ventas_entregadas,
        # Elimina las variables de paginación del context si no usas paginación
    }

    return render(request, 'admin/ventas/ventas_admin.html', context)
from django.shortcuts import render, redirect
from django.forms import DurationField, modelformset_factory
from .models import Venta, DetalleVenta, Pago, CombinacionProducto
from .forms import VentaForm, DetalleVentaForm, PagoForm, CombinacionProductoForm, CombinacionProductoFormSet, DetalleVentaFormSet
from django.db import transaction
from  decimal import Decimal
from django.contrib import messages


@login_required
@transaction.atomic
def agregar_venta_completa(request):
    if request.method == 'POST':
        venta_form = VentaForm(request.POST)
        detalle_formset = DetalleVentaFormSet(request.POST, prefix='detalle')
        combinacion_formset = CombinacionProductoFormSet(request.POST, prefix='combinacion')
        pago_form = PagoForm(request.POST)

        if venta_form.is_valid() and detalle_formset.is_valid() and combinacion_formset.is_valid() and pago_form.is_valid():
            # 1. Calcular cantidad total de donas
            total_donas = sum(
                form.cleaned_data['cantidad']
                for form in detalle_formset
                if form.cleaned_data and not form.cleaned_data.get('DELETE', False)
            )
            total_combinaciones = sum(
                1 for form in combinacion_formset
                if form.cleaned_data and not form.cleaned_data.get('DELETE', False)
            )

            if total_donas != total_combinaciones:
                messages.error(request, f'Debes agregar exactamente {total_donas} combinaciones. Has agregado {total_combinaciones}.')
            else:
                try:
                    with transaction.atomic():
                        venta = venta_form.save(commit=False)
                        subtotal = Decimal('0.00')
                        detalles_guardados = []

                        # Primero guardar la venta para tener la PK
                        venta.subtotal = 0
                        venta.iva = 0
                        venta.total = 0
                        venta.save()

                        # Guardar detalles con subtotal calculado manualmente
                        for form in detalle_formset:
                            if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                                detalle = form.save(commit=False)
                                detalle.cod_venta = venta
                                detalle.save()
                                detalles_guardados.append({
                                    'detalle': detalle,
                                    'cantidad': detalle.cantidad
                                })
                                subtotal += detalle.cantidad * detalle.precio_unitario  # 👈 usa fórmula

                        # Calcular totales
                        iva = subtotal * Decimal('0.19')
                        total = subtotal + iva

                        venta.subtotal = subtotal
                        venta.iva = iva
                        venta.total = total
                        venta.save()

                        # Guardar combinaciones en orden
                        # Guardar combinaciones en orden
                        combinacion_index = 0
                        for d in detalles_guardados:
                            for _ in range(d['cantidad']):
                                form = combinacion_formset.forms[combinacion_index]
                                if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                                    combinacion = form.save(commit=False)
                                    combinacion.cod_detalle = d['detalle']
                                    combinacion.cod_producto = d['detalle'].cod_producto  # ← AGREGAR ESTO
                                    combinacion.save()
                                combinacion_index += 1


                        # Guardar pago
                        pago = pago_form.save(commit=False)
                        pago.cod_venta = venta
                        pago.save()

                        messages.success(request, "Venta registrada correctamente.")
                        return redirect('ventas_admin')
                except Exception as e:
                    messages.error(request, f"Ocurrió un error al guardar la venta: {e}")
        else:
            messages.error(request, "Por favor corrige los errores en el formulario.")
    else:
        venta_form = VentaForm()
        detalle_formset = DetalleVentaFormSet(prefix='detalle', queryset=DetalleVenta.objects.none(), initial=[{}])
        combinacion_formset = CombinacionProductoFormSet(prefix='combinacion', queryset=CombinacionProducto.objects.none(), initial=[{}])
        pago_form = PagoForm()

    productos = Producto.objects.filter(activo=True)
    precios = {str(p.cod_producto): str(p.prec_pro) for p in productos}

    return render(request, 'admin/ventas/agregar_venta_completa.html', {
        'venta_form': venta_form,
        'detalle_formset': detalle_formset,
        'combinacion_formset': combinacion_formset,
        'pago_form': pago_form,
        'productos': productos,
        'precios_productos': precios,
    })

@login_required
def detalle_ventas(request, venta_id):
    venta = get_object_or_404(Venta, pk=venta_id)

    # Traer detalles con producto relacionado
    detalles = DetalleVenta.objects.filter(cod_venta=venta).select_related('cod_producto')

    total = sum(detalle.subtotal for detalle in detalles)

    context = {
        'venta': venta,
        'detalles': detalles,
        'total': total,
    }
    return render(request, 'admin/ventas/detalle_ventas.html', context)

@login_required
def editar_estado_venta(request, venta_id):
    venta = get_object_or_404(Venta, pk=venta_id)

    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        if nuevo_estado in dict(Venta.ESTADOS).keys():
            venta.estado = nuevo_estado
            venta.save()
            return redirect('ventas_admin')

    return render(request, 'admin/ventas/editar_estado_venta.html', {
        'venta': venta,
        'estados': Venta.ESTADOS
    })


# Vista de Producción
from django.shortcuts import render, get_object_or_404, redirect
from .models import Produccion, Salida, Entrada, Insumo, Venta, Envio
from .forms import ProduccionForm, SalidaForm, EntradaForm, EnvioForm
from django.db import transaction

@login_required
def produccion_admin(request):
    producciones = Produccion.objects.select_related('cod_venta').all()
    envios = Envio.objects.all()

    # Diccionario para saber si ya existe un envío para una venta específica
    envios_por_venta = {envio.cod_venta_id: envio for envio in envios}

    context = {
        'producciones': producciones,
        'envios_por_venta': envios_por_venta,
    }
    return render(request, 'admin/produccion/produccion_admin.html', context)


from django.db import transaction
from django.utils import timezone
from datetime import datetime
def confirmar_generacion_produccion(request, cod_venta):
    venta = get_object_or_404(Venta, cod_venta=cod_venta)
    detalles = DetalleVenta.objects.filter(cod_venta=venta)

    insumos_requeridos = {}

    for detalle in detalles:
        producto = detalle.cod_producto
        cantidad = detalle.cantidad
        multiplicador = producto.get_multiplicador()

        # 1. Insumos de receta base
        receta = RecetaProducto.objects.filter(cod_producto=producto)
        for r in receta:
            insumo = r.insumo
            key = insumo.nomb_insumo
            insumos_requeridos.setdefault(key, {
                "cantidad": 0,
                "stock": insumo.cnt_insumo,
                "unidad": insumo.unidad_medida,
                "fuente": set()
            })
            insumos_requeridos[key]["cantidad"] += r.cantidad * cantidad
            insumos_requeridos[key]["fuente"].add("Receta base")

        # 2. Insumos por personalización
        combinaciones = CombinacionProducto.objects.filter(cod_detalle=detalle)
        for c in combinaciones:
            combinaciones_insumos = [
                (c.cod_sabor_masa_1, 15),     # 15 ml de masa base
                (c.cod_glaseado_1, 200),      # 200 g de glaseado
                (c.cod_topping_1, 100),       # 100 g de topping
            ]
            for componente, cantidad_base in combinaciones_insumos:
                if componente and componente.insumo:
                    insumo = componente.insumo
                    key = insumo.nomb_insumo
                    insumos_requeridos.setdefault(key, {
                        "cantidad": 0,
                        "stock": insumo.cnt_insumo,
                        "unidad": insumo.unidad_medida,
                        "fuente": set()
                    })
                    total_personalizado = cantidad * multiplicador * cantidad_base
                    insumos_requeridos[key]["cantidad"] += total_personalizado
                    insumos_requeridos[key]["fuente"].add("Personalización")

    # Verificar si hay stock suficiente
    hay_stock_suficiente = all(data["cantidad"] <= data["stock"] for data in insumos_requeridos.values())

    context = {
        'venta': venta,
        'insumos_requeridos': insumos_requeridos,
        'hay_stock_suficiente': hay_stock_suficiente,
    }
    return render(request, 'admin/ventas/confirmar_generacion_produccion.html', context)

from django.forms import modelform_factory
from django.forms import formset_factory
from django.forms import modelform_factory
from django.forms import formset_factory

def seleccionar_combinaciones_admin(request, cod_venta):
    venta = get_object_or_404(Venta, cod_venta=cod_venta)
    detalles = DetalleVenta.objects.filter(cod_venta=venta)

    masas = SaborMasa.objects.all()
    glaseados = Glaseado.objects.all()
    toppings = Topping.objects.all()

    if request.method == 'POST':
        combinaciones = []

        index = 0
        for detalle in detalles:
            for i in range(detalle.cantidad):
                masa_id = request.POST.get(f'masa_{index}')
                glaseado_id = request.POST.get(f'glaseado_{index}')
                topping_id = request.POST.get(f'topping_{index}')

                if not (masa_id and glaseado_id and topping_id):
                    messages.error(request, "Debes seleccionar todas las combinaciones.")
                    return redirect(request.path)

                combinaciones.append({
                    'detalle_id': detalle.id,
                    'masa_id': masa_id,
                    'glaseado_id': glaseado_id,
                    'topping_id': topping_id,
                })
                index += 1

        request.session['combinaciones'] = combinaciones
        return redirect('generar_produccion_admin', cod_venta=venta.cod_venta)

    unidades = []
    for detalle in detalles:
        for i in range(detalle.cantidad):
            unidades.append({
                'detalle': detalle,
                'numero': i + 1,
            })

    context = {
        'venta': venta,
        'unidades': unidades,
        'masas': masas,
        'glaseados': glaseados,
        'toppings': toppings,
    }
    return render(request, 'admin/ventas/seleccionar_combinaciones.html', context)


CANTIDAD_BASE = {
    'masa': 15,
    'glaseado': 200,
    'topping': 100,
}

@transaction.atomic
def generar_produccion_confirmada(request, cod_venta):
    venta = get_object_or_404(Venta, cod_venta=cod_venta)

    if request.method == 'POST':
        if hasattr(venta, 'produccion'):
            messages.warning(request, "Ya se ha generado una producción para esta venta.")
            return redirect('ventas_admin')

        produccion = Produccion.objects.create(
            cod_venta=venta,
            fecha_inicio=datetime.now()
        )

        insumos_necesarios = {}

        for detalle in DetalleVenta.objects.filter(cod_venta=venta):
            producto = detalle.cod_producto
            multiplicador = producto.get_multiplicador()

            # --- 1. Receta base ---
            receta = RecetaProducto.objects.filter(cod_producto=producto)
            for item in receta:
                total = item.cantidad * detalle.cantidad
                insumos_necesarios[item.insumo_id] = insumos_necesarios.get(item.insumo_id, 0) + total

            # --- 2. Personalización por combinación ---
            combinaciones = CombinacionProducto.objects.filter(cod_detalle=detalle)
            for combinacion in combinaciones:
                # Masa
                if combinacion.cod_sabor_masa_1 and combinacion.cod_sabor_masa_1.insumo_id:
                    id_insumo = combinacion.cod_sabor_masa_1.insumo_id
                    cantidad = CANTIDAD_BASE['masa'] * multiplicador
                    insumos_necesarios[id_insumo] = insumos_necesarios.get(id_insumo, 0) + cantidad

                # Glaseado
                if combinacion.cod_glaseado_1 and combinacion.cod_glaseado_1.insumo_id:
                    id_insumo = combinacion.cod_glaseado_1.insumo_id
                    cantidad = CANTIDAD_BASE['glaseado'] * multiplicador
                    insumos_necesarios[id_insumo] = insumos_necesarios.get(id_insumo, 0) + cantidad

                # Topping
                if combinacion.cod_topping_1 and combinacion.cod_topping_1.insumo_id:
                    id_insumo = combinacion.cod_topping_1.insumo_id
                    cantidad = CANTIDAD_BASE['topping'] * multiplicador
                    insumos_necesarios[id_insumo] = insumos_necesarios.get(id_insumo, 0) + cantidad

        # Verificación de stock
        faltantes = []
        for insumo_id, requerido in insumos_necesarios.items():
            try:
                insumo = Insumo.objects.get(pk=insumo_id)
            except Insumo.DoesNotExist:
                messages.error(request, f"No existe el insumo con ID {insumo_id}.")
                return redirect('ventas_admin')
            if insumo.cnt_insumo < requerido:
                faltantes.append(f"{insumo.nomb_insumo} (faltan {requerido - insumo.cnt_insumo:.2f})")

        if faltantes:
            messages.error(request, "No hay suficiente stock para:\n" + "\n".join(faltantes))
            raise transaction.TransactionManagementError  # Revierte todo si hay faltantes

        # Registrar salidas y descontar stock
        for insumo_id, cantidad in insumos_necesarios.items():
            insumo = Insumo.objects.get(pk=insumo_id)
            Salida.objects.create(
                cod_produccion=produccion,
                cod_insumo=insumo,
                cantidad=cantidad
            )
            insumo.cnt_insumo -= cantidad
            insumo.save()

        messages.success(request, "Producción generada exitosamente y stock actualizado.")
        return redirect('produccion_admin')

    messages.error(request, "Acceso inválido.")
    return redirect('ventas_admin')


def editar_produccion(request, cod_produccion):
    produccion = get_object_or_404(Produccion, pk=cod_produccion)

    # Si ya tiene un envío asignado, no se permite editar
    if produccion.cod_envio:
        messages.warning(request, "Esta producción ya tiene un envío asignado y no se puede volver a asignar.")
        return redirect('produccion_admin')

    if request.method == 'POST':
        form = ProduccionForm(request.POST, instance=produccion)
        if form.is_valid():
            form.save()
            messages.success(request, "Producción actualizada correctamente.")
            return redirect('produccion_admin')
    else:
        form = ProduccionForm(instance=produccion)

    return render(request, 'admin/produccion/editar_produccion.html', {
        'form': form,
        'produccion': produccion
    })
def cambiar_estado_produccion(request, cod_produccion):
    produccion = get_object_or_404(Produccion, pk=cod_produccion)

    if request.method == 'POST':
        nuevo = request.POST.get('estado')
        if nuevo in dict(Produccion.ESTADOS):
            produccion.estado = nuevo
            if nuevo == 'FINALIZADO':
                produccion.fecha_fin = timezone.now()
            produccion.save()

            # Mover esto aquí para que se ejecute al cambiar el estado
            venta = produccion.cod_venta
            if nuevo == 'EN_PROCESO':
                venta.estado = 'PREPARACION'
            elif nuevo == 'FINALIZADO':
                venta.estado = 'EN_CAMINO'
            elif nuevo == 'PENDIENTE':
                venta.estado = 'PENDIENTE'
            venta.save()

        return redirect('produccion_admin')

    return render(request, 'admin/produccion/cambiar_estado_produccion.html', {
        'produccion': produccion,
        'estados': Produccion.ESTADOS
    })


@login_required
def asignar_envio_produccion(request, cod_produccion):
    produccion = get_object_or_404(Produccion, pk=cod_produccion)
    venta = produccion.cod_venta  # Relación desde Producción a Venta

    # 🔐 Verificar si ya existe un envío para esta venta
    if Envio.objects.filter(cod_venta=venta).exists():
        messages.error(request, "Ya existe un envío asignado para esta venta.")
        return redirect('produccion_admin')

    if request.method == 'POST':
        form = EnvioForm(request.POST)
        if form.is_valid():
            envio = form.save(commit=False)
            envio.cod_venta = venta
            if not envio.estado or envio.estado == 'PENDIENTE':
                envio.estado = 'ASIGNADO'
            envio.save()

            # Cambiar el estado de la venta a EN_CAMINO
            venta.estado = 'EN_CAMINO'
            venta.save()
            
            messages.success(request, "¡Envío asignado correctamente!")
            return redirect('produccion_admin')
        else:
            messages.error(request, "Hay errores en el formulario. Revisa los campos.")
    else:
        form = EnvioForm()

    return render(request, 'admin/produccion/asignar_envio_produccion.html', {
        'form': form,
        'venta': venta,
        'produccion': produccion
    })

def eliminar_produccion(request, cod_produccion):
    produccion = get_object_or_404(Produccion, pk=cod_produccion)
    if request.method == 'POST':
        produccion.delete()
        return redirect('produccion_admin')
    return render(request, 'admin/produccion/eliminar_produccion.html', {'produccion': produccion})

#Envio Admin
from django.shortcuts import render, redirect, get_object_or_404
from .models import Envio
from .forms import EnvioForm

# Vista de Envíos
@login_required
def envios_admin(request):
    envios = Envio.objects.all()
    return render(request, 'admin/envios/envios_admin.html', {'envios': envios})

def crear_envio(request):
    if request.method == 'POST':
        form = EnvioForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('envios_admin')
    else:
        form = EnvioForm()
    return render(request, 'admin/envios/crear_envio.html', {'form': form, 'titulo': 'Asignar Envío'})

def editar_envio(request, pk):
    envio = get_object_or_404(Envio, pk=pk)
    if request.method == 'POST':
        form = EnvioForm(request.POST, instance=envio)
        if form.is_valid():
            form.save()
            return redirect('envios_admin')
    else:
        form = EnvioForm(instance=envio)
    return render(request, 'admin/envios/editar_envio.html', {'form': form, 'titulo': 'Editar Envío'})

def eliminar_envio(request, pk):
    envio = get_object_or_404(Envio, pk=pk)
    envio.delete()
    return redirect('envios_admin')






# Vista de Proveedores
from .models import Proveedor
from .forms import ProveedorForm
@login_required
def proveedores_admin(request):
    proveedores = Proveedor.objects.all()
    return render(request, 'admin/proveedores/proveedores_admin.html', {'proveedores': proveedores})


def agregar_proveedores(request):
    if request.method == 'POST':
        Proveedor.objects.create(
            nom_proveedor=request.POST['nom_proveedor'],
            telefono_proveedor=request.POST['telefono_proveedor'],
            direccion_proveedor=request.POST.get('direccion_proveedor', ''),
            email_proveedor=request.POST.get('email_proveedor', ''),
            novedad_proveedor=request.POST.get('novedad_proveedor', '')
        )
        return redirect('proveedores_admin')
    return render(request, 'admin/proveedores/agregar_proveedores.html')
    
def editar_proveedores(request, cod_proveedor):
    proveedor = get_object_or_404(Proveedor, pk=cod_proveedor)
    if request.method == 'POST':
        proveedor.nom_proveedor = request.POST['nom_proveedor']
        proveedor.telefono_proveedor = request.POST['telefono_proveedor']
        proveedor.direccion_proveedor = request.POST.get('direccion_proveedor', '')
        proveedor.email_proveedor = request.POST.get('email_proveedor', '')
        proveedor.novedad_proveedor = request.POST.get('novedad_proveedor', '')
        proveedor.save()
        return redirect('proveedores_admin')
    return render(request, 'admin/proveedores/editar_proveedores.html', {'proveedor': proveedor})


def eliminar_proveedores(request, cod_proveedor):
    proveedor = get_object_or_404(Proveedor, pk=cod_proveedor)
    if request.method == 'POST':
        proveedor.delete()
        return redirect('proveedores_admin')
    return render(request, 'admin/proveedores/eliminar_proveedores.html', {'proveedor': proveedor})




# Vista de Entradas (Inventario)
from django.shortcuts import render, get_object_or_404, redirect
from .models import Entrada
from .forms import EntradaForm

@login_required
def entradas_admin(request):
    entradas = Entrada.objects.all().order_by('-fecha_hora_entrada')
    return render(request, 'admin/entradas/entradas_admin.html', {'entradas': entradas})


# Agregar entrada
def agregar_entradas(request):
    if request.method == 'POST':
        form = EntradaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('entradas_admin')  # Redirige a la lista de entradas
    else:
        form = EntradaForm()
    return render(request, 'admin/entradas/agregar_entradas.html', {'form': form})


# Editar entrada
def editar_entrada(request, cod_entrada):
    entrada = get_object_or_404(Entrada, cod_entrada=cod_entrada)
    insumo_anterior = entrada.cod_insumo
    cantidad_anterior = entrada.cnt_entrada

    if request.method == 'POST':
        form = EntradaForm(request.POST, instance=entrada)
        if form.is_valid():
            nueva_entrada = form.save(commit=False)
            nuevo_insumo = nueva_entrada.cod_insumo
            nueva_cantidad = nueva_entrada.cnt_entrada

            if insumo_anterior == nuevo_insumo:
                # Mismo insumo: ajustar diferencia de cantidad
                diferencia = nueva_cantidad - cantidad_anterior
                nuevo_insumo.cnt_insumo += diferencia
                nuevo_insumo.save()
            else:
                # Cambió el insumo: revertir en el anterior y sumar en el nuevo
                insumo_anterior.cnt_insumo -= cantidad_anterior
                insumo_anterior.save()

                nuevo_insumo.cnt_insumo += nueva_cantidad
                nuevo_insumo.save()

            nueva_entrada.save()
            return redirect('entradas_admin')
    else:
        form = EntradaForm(instance=entrada)

    return render(request, 'admin/entradas/editar_entradas.html', {'form': form})


# Eliminar entrada
def eliminar_entrada(request, cod_entrada):
    entrada = get_object_or_404(Entrada, cod_entrada=cod_entrada)
    if request.method == 'POST':
        # Descontar del stock
        entrada.cod_insumo.cnt_insumo -= entrada.cnt_entrada
        entrada.cod_insumo.save()

        entrada.delete()
        return redirect('entradas_admin')  # Redirige a la lista de entradas
    return render(request, 'admin/entradas/eliminar_entradas.html', {'entrada': entrada})






# Vista de Salidas (Inventario)
@login_required
def salidas_admin(request):
    # ✅ Corregido: usar el modelo 'Salida' directamente
    salidas = Salida.objects.all().order_by('-fecha_hora_salida')
    return render(request, 'admin/salidas/salidas_admin.html', {'salidas': salidas})
def eliminar_salida(request, cod_salida):
    salida = get_object_or_404(Salida, cod_salida=cod_salida)
    salida.delete()
    return redirect('salidas_admin')  # Redirige a la lista de salidas


def agregar_salida(request):
    if request.method == 'POST':
        form = SalidaForm(request.POST)
        if form.is_valid():
            salida = form.save()

            # Descontar del inventario del insumo
            insumo = salida.cod_insumo
            insumo.cnt_insumo -= salida.cantidad
            insumo.save()

            return redirect('salidas_admin')  # Redirige a la lista de salidas
    else:
        form = SalidaForm()
    
    return render(request, 'admin/salidas/agregar_salida.html', {'form': form})


# Vista de Categorías
from .models import CategoriaInsumo
@login_required
def categorias_admin(request):
    # ✅ Corregido: usar 'CategoriaInsumo' en lugar de 'Categoria'
    categorias = CategoriaInsumo.objects.all()
    return render(request, 'admin/categorias/categorias_admin.html', {'categorias': categorias})


def agregar_categoria(request):
    if request.method == 'POST':
        nom_categoria = request.POST.get('nom_categoria').strip()
        descripcion = request.POST.get('descripcion', '').strip()

        if CategoriaInsumo.objects.filter(nom_categoria__iexact=nom_categoria).exists():
            messages.warning(request, 'Ya existe una categoría con ese nombre.')
            return redirect('agregar_categoria')
        
        CategoriaInsumo.objects.create(
            nom_categoria=nom_categoria,
            descripcion=descripcion or None
        )
        messages.success(request, 'Categoría agregada exitosamente.')
        return redirect('categorias_admin')

    return render(request, 'admin/categorias/agregar_categoria.html')


def eliminar_categoria(request, cod_categoria):
    categoria = get_object_or_404(CategoriaInsumo, cod_categoria=cod_categoria)
    
    try:
        categoria.delete()
        messages.success(request, 'Categoría eliminada correctamente.')
    except:
        messages.error(request, 'No se pudo eliminar la categoría. Puede estar relacionada con insumos.')

    return redirect('categorias_admin')



# Vista de Correos - COMENTADA porque no tienes este modelo
# Si necesitas esta funcionalidad, debes crear el modelo Correo

from .models import Correo

@login_required
def correos_admin(request):
    historial_correos = Correo.objects.all().order_by('-fecha_envio')
    return render(request, 'admin/correos/correos_admin.html', {
        'historial_correos': historial_correos
    })

def cargar_insumos(request):
    if request.method == 'POST':
        archivo = request.FILES.get('archivo_insumos')
        if not archivo or not archivo.name.endswith('.csv'):
            messages.error(request, 'El archivo debe ser un CSV.')
            return redirect('cargarDatos')
        try:
            decoded_file = archivo.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            insumos_cargados = 0
            insumos_fallidos = []
            for fila in reader:
                # Ajusta los nombres de las columnas según tu CSV
                nomb_insumo = fila.get('nomb_insumo')
                cnt_insumo = fila.get('cnt_insumo')
                unidad_medida = fila.get('unidad_medida')
                cod_categoria_id = fila.get('cod_categoria') or fila.get('cod_categoria_id')
                fecha_vencimiento = fila.get('fecha_vencimiento')

                categoria = None
                if cod_categoria_id:
                    try:
                        categoria = CategoriaInsumo.objects.get(pk=cod_categoria_id)
                    except CategoriaInsumo.DoesNotExist:
                        categoria = None

                if categoria and nomb_insumo and cnt_insumo and unidad_medida and fecha_vencimiento:
                    try:
                        Insumo.objects.create(
                            nomb_insumo=nomb_insumo,
                            cnt_insumo=int(cnt_insumo),
                            unidad_medida=unidad_medida,
                            cod_categoria=categoria,
                            fecha_vencimiento=fecha_vencimiento
                        )
                        insumos_cargados += 1
                    except Exception as e:
                        insumos_fallidos.append(f"{nomb_insumo} (error: {e})")
                else:
                    insumos_fallidos.append(nomb_insumo or 'Sin nombre')

            msg = f'Insumos cargados exitosamente: {insumos_cargados}.'
            if insumos_fallidos:
                msg += f' No se cargaron estos insumos: {", ".join(insumos_fallidos)}.'
                messages.warning(request, msg)
            else:
                messages.success(request, msg)
        except Exception as e:
            messages.error(request, f'Error al cargar insumos: {e}')
        return redirect('cargarDatos')
    return redirect('cargarDatos')
    
def ver_correo(request, cod_correo):
    correo = get_object_or_404(Correo, cod_correo=cod_correo)
    return render(request, 'admin/correos/ver_correo.html', {'correo': correo})
  
from django.core.mail import EmailMessage
from .models import Correo
from .models import Cliente, Administrador, Domiciliario  # Ajusta si usas otros nombres

def enviar_correos_masivos(request):
    if request.method == 'POST':
        destinatarios_tipo = request.POST.getlist('destinatarios')
        asunto = request.POST.get('asunto')
        mensaje = request.POST.get('mensaje')
        archivos = request.FILES.getlist('adjuntos')

        # Obtener correos reales de los destinatarios seleccionados
        correos_a_enviar = []

        if 'CLIENTE' in destinatarios_tipo:
            correos_a_enviar += list(Cliente.objects.select_related('cod_usua').values_list('cod_usua__email', flat=True))

        if 'ADMIN' in destinatarios_tipo:
            correos_a_enviar += list(Administrador.objects.select_related('cod_usua').values_list('cod_usua__email', flat=True))

        if 'DOMI' in destinatarios_tipo:
            correos_a_enviar += list(Domiciliario.objects.select_related('cod_usua').values_list('cod_usua__email', flat=True))

        # Enviar correos y registrar en BD
        enviados = 0
        for correo_dest in set(correos_a_enviar):  # Evitar duplicados
            try:
                email = EmailMessage(
                    asunto,
                    mensaje,
                    to=[correo_dest]
                )

                # Adjuntar archivos
                for adj in archivos:
                    email.attach(adj.name, adj.read(), adj.content_type)

                email.send()
                enviado = True
                enviados += 1
            except Exception as e:
                enviado = False  # puedes registrar el error si lo deseas

            # Guardar registro en BD
            Correo.objects.create(
                destinatario=correo_dest,
                asunto=asunto,
                mensaje=mensaje,
                enviado=enviado
            )

        messages.success(request, f'Se enviaron {enviados} correos.')
        return redirect('correos_admin')

    return redirect('admin/correos/correos_admin')


# Vista para Cargar Datos
@login_required
def cargarDatos(request):
    if request.method == 'POST':
        form = CargarDatosForm(request.POST, request.FILES)
        if form.is_valid():
            # Procesar el archivo subido
            archivo = request.FILES['archivo']
            # Aquí iría tu lógica para procesar el archivo
            return redirect('dashboard')
    else:
        form = CargarDatosForm()
    
    return render(request, 'admin/cargarDatos.html', {'form': form})

# 🔥 VISTAS ADICIONALES QUE PODRÍAS NECESITAR

from django.shortcuts import render, redirect, get_object_or_404
from .models import Insumo
from .forms import InsumoForm  # Asegúrate de tener este formulario creado
from django.db.models import Q
from django.utils import timezone

@login_required
def insumos_admin(request):
    """Vista para gestionar insumos"""
    from .models import Insumo
    insumos = Insumo.objects.select_related('cod_categoria').all()

    mes_filtro = request.GET.get("mes")
    if mes_filtro:
        insumos = insumos.filter(
            fecha_vencimiento__month=int(mes_filtro)
        )

    meses_lista = [
        ("01", "Enero"), ("02", "Febrero"), ("03", "Marzo"), ("04", "Abril"),
        ("05", "Mayo"), ("06", "Junio"), ("07", "Julio"), ("08", "Agosto"),
        ("09", "Septiembre"), ("10", "Octubre"), ("11", "Noviembre"), ("12", "Diciembre")
    ]

    return render(request, 'admin/insumos/insumos_admin.html', {
        'insumos': insumos,
        'meses_lista': meses_lista,
    })
    

# Agregar insumo
def agregar_insumo(request):
    if request.method == 'POST':
        form = InsumoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('insumos_admin')
    else:
        form = InsumoForm()
    return render(request, 'admin/insumos/agregar_insumo.html', {'form': form})

# Editar insumo
def editar_insumo(request, cod_insumo):
    insumo = get_object_or_404(Insumo, cod_insumo=cod_insumo)
    if request.method == 'POST':
        form = InsumoForm(request.POST, instance=insumo)
        if form.is_valid():
            form.save()
            return redirect('insumos_admin')
    else:
        form = InsumoForm(instance=insumo)
    return render(request, 'admin/insumos/editar_insumo.html', {'form': form})

# Eliminar insumo
def eliminar_insumo(request, cod_insumo):
    insumo = get_object_or_404(Insumo, cod_insumo=cod_insumo)
    if request.method == 'POST':
        insumo.delete()
        # Si es AJAX/fetch, responde con JSON
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.META.get('HTTP_ACCEPT') == 'application/json':
            return JsonResponse({'success': True})
        return redirect('insumos_admin')
    # Si GET, muestra confirmación (opcional)
    return render(request, 'admin/insumos/eliminar_insumo.html', {'insumo': insumo})



@login_required
def clientes_admin(request):
    """Vista para gestionar clientes"""
    from .models import Cliente
    clientes = Cliente.objects.select_related('cod_usua').all()
    return render(request, 'admin/clientes_admin.html', {'clientes': clientes})

@login_required
def domiciliarios_admin(request):
    """Vista para gestionar domiciliarios"""
    from .models import Domiciliario
    domiciliarios = Domiciliario.objects.select_related('cod_usua').all()
    return render(request, 'admin/domiciliarios_admin.html', {'domiciliarios': domiciliarios})

    
def actualizar_usuario(request):
    if request.method == 'POST':
        cod_usuario = request.POST.get('codUsuario')

        nombre = request.POST.get('nomUsua')
        apellido = request.POST.get('apellUsua')
        email = request.POST.get('emailUsua')
        rol = request.POST.get('rol')

        usuario = get_object_or_404(Usuario, pk=cod_usuario)
        usuario.nom_usua = nombre
        usuario.apell_usua = apellido
        usuario.email_usua = email
        usuario.rol = rol
        usuario.save()

        return redirect(request.META.get('HTTP_REFERER', '/'))  # Redirige a la misma página
    
def eliminar_usuario(request, cod_usuario):
      usuario = get_object_or_404(Usuario, pk=cod_usuario)
      usuario.delete()
      return redirect(request.META.get('HTTP_REFERER', '/'))
      return redirect(request.META.get('HTTP_REFERER', '/'))


from django.contrib.auth.hashers import make_password
import csv

def cargar_datos_view(request):
    if request.method == 'POST':
        archivo = request.FILES.get('archivo')

        # Validar extensión
        if not archivo.name.endswith('.csv'):
            messages.error(request, 'El archivo debe tener extensión .csv')
            return redirect('cargarDatos')

        try:
            decoded_file = archivo.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)

            for fila in reader:
                Usuario.objects.update_or_create(
                    email=fila['email'],
                    defaults={
                        'nom_usua': fila['nombre'],
                        'apell_usua': fila['apellido'],
                        'password': make_password(fila['contraseña']),
                        'rol': fila['rol'].upper(),  # Asegura que el rol esté en mayúsculas
                    }
                )

            messages.success(request, 'Datos cargados exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {e}')

        return redirect('cargarDatos')  # Redirige para recargar la página y mostrar mensajes

    return render(request, 'admin/cargarDatos.html')  # Reemplaza con tu template real


#productos
from django.shortcuts import render, redirect, get_object_or_404
from .models import Producto, RecetaProducto, Insumo
from .forms import RecetaProductoFormSet # Lo creamos más abajo
from django.forms import modelformset_factory
from .forms import RecetaProducto

def productos_admin(request):
    productos = Producto.objects.all().order_by('activo', 'tamano')
    return render(request, 'admin/productos/productos_admin.html', {'productos': productos})

def cambiar_estado_producto(request, cod_producto):
    producto = get_object_or_404(Producto, pk=cod_producto)
    producto.activo = not producto.activo
    producto.save()
    return redirect('productos_admin')

from .forms import RecetaProductoFormSet

def ver_receta_producto(request, cod_producto):
    producto = get_object_or_404(Producto, cod_producto=cod_producto)
    receta = RecetaProducto.objects.filter(cod_producto=producto)
    return render(request, 'admin/productos/ver_receta.html', {'producto': producto, 'receta': receta})





INSUMOS_BASICOS = [

    'Huevos',
    'Azucar',
    'Sal',
    'Mantequilla',
    'Leche',
    'Harina de trigo',
    'Polvo para hornear',
]


def editar_receta_producto(request, cod_producto):
    producto = get_object_or_404(Producto, cod_producto=cod_producto)

    # Obtener los insumos básicos desde la base de datos
    insumos_basicos = Insumo.objects.filter(nomb_insumo__in=INSUMOS_BASICOS)

    # Crear receta vacía si aún no existe
    for insumo in insumos_basicos:
        RecetaProducto.objects.get_or_create(
            cod_producto=producto,
            insumo=insumo,
            defaults={
                'cantidad': 0,
                'unidad_medida': 'g',  # o ml, unidades, según el insumo
            }
        )

    receta_qs = RecetaProducto.objects.filter(cod_producto=producto)

    if request.method == 'POST':
        formset = RecetaProductoFormSet(request.POST, queryset=receta_qs)
        if formset.is_valid():
            formset.save()
            return redirect('productos_admin')
    else:
        formset = RecetaProductoFormSet(queryset=receta_qs)

    return render(request, 'admin/productos/editar_receta_producto.html', {
        'producto': producto,
        'formset': formset
    })

from django.shortcuts import render, redirect, get_object_or_404
from .models import Producto, RecetaPersonalizacion
from .forms import RecetaPersonalizacionFormSet
from django.contrib import messages

def editar_receta_personalizada(request, cod_producto):
    producto = get_object_or_404(Producto, cod_producto=cod_producto)
    queryset = RecetaPersonalizacion.objects.filter(cod_producto=producto)

    if request.method == 'POST':
        formset = RecetaPersonalizacionFormSet(request.POST, queryset=queryset)
        if formset.is_valid():
            recetas = formset.save(commit=False)
            # Eliminar los que el usuario marcó para borrar
            for obj in formset.deleted_objects:
                obj.delete()
            for receta in recetas:
                receta.cod_producto = producto
                # Asignar cantidad base automáticamente según tipo
                if receta.tipo == 'MASA':
                    receta.cantidad_base = 15  # gramos (1 cucharada)
                elif receta.tipo == 'GLASEADO':
                    receta.cantidad_base = 200
                elif receta.tipo == 'TOPPING':
                    receta.cantidad_base = 100
                receta.save()
            messages.success(request, "Recetas personalizadas guardadas correctamente.")
            return redirect('editar_receta_personalizada', cod_producto=producto.cod_producto)
    else:
        formset = RecetaPersonalizacionFormSet(queryset=queryset)

    return render(request, 'admin/productos/editar_receta_personalizada.html', {
        'formset': formset,
        'producto': producto,
    })


def generar_receta_base(request, cod_producto):

    producto = get_object_or_404(Producto, cod_producto=cod_producto)

    # Elimina receta anterior si existe
    RecetaProducto.objects.filter(cod_producto=producto).delete()

    # Ingredientes base para S
    ingredientes_base = [
        ('Azucar', 50, 'Gramos'),
        ('Sal', 1, 'Gramos'),
        ('Huevos', 1, 'Unidad'),
        ('Mantequilla', 15, 'Gramos'),
        ('Leche', 100, 'Litros'),
        ('Harina de trigo', 100, 'Gramos'),
        ('Polvo para hornear', 5, 'Gramos'),
    ]

    tamano = producto.tamano.upper()
    multiplicador = {'S': 1, 'M': 2, 'L': 3, 'XL': 4}.get(tamano, 1)

    for nombre_insumo, cantidad_base, unidad in ingredientes_base:
        try:
            insumo = Insumo.objects.get(nomb_insumo__iexact=nombre_insumo)
            RecetaProducto.objects.create(
                cod_producto=producto,
                insumo=insumo,
                cantidad=cantidad_base * multiplicador,
                unidad_medida=unidad
            )
        except Insumo.DoesNotExist:
            messages.error(request, f"No se encontró el insumo '{nombre_insumo}'.")

    messages.success(request, "Receta base generada correctamente.")
    return redirect('editar_receta_producto', cod_producto=producto.cod_producto)






def reporte_usuarios_pdf(request):
    # Estadísticas: contar cuántos usuarios hay por rol
   
    roles = ['ADMIN', 'CLIENTE', 'DOMI']
    conteo_roles = [Usuario.objects.filter(rol=rol).count() for rol in roles]

    # Generar el gráfico
    plt.figure(figsize=(6, 4))  # tamaño opcional
    plt.bar(roles, conteo_roles, color='skyblue')
    plt.title('Usuarios por Rol')
    plt.xlabel('Rol')
    plt.ylabel('Cantidad')
    plt.tight_layout()

    # Guardar gráfico en memoria
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    grafico_base64 = base64.b64encode(buf.read()).decode('utf-8')
    buf.close()

    # Cerrar el plot (importantísimo para evitar sobrecarga)
    plt.close()

    # Renderizar template HTML con gráfico embebido
    html_string = render_to_string('reportes/usuarios_pdf.html', {
        'grafico_base64': grafico_base64,
        'total_admin': conteo_roles[0],
        'total_cliente': conteo_roles[1],
        'total_domi': conteo_roles[2],
    })

    html = HTML(string=html_string)
    pdf_file = html.write_pdf()

    # Devolver PDF como respuesta
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="reporte_usuarios.pdf"'
    return response

import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from .models import Usuario

def reporte_usuarios_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    # Cabeceras
    encabezados = ['ID', 'Nombre', 'Apellido', 'Email', 'Rol', 'Activo']
    ws.append(encabezados)

    # Formato de cabeceras
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(encabezados)):
        for cell in col:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    # Llenar filas con los datos
    usuarios = Usuario.objects.all()
    for usuario in usuarios:
        ws.append([
            usuario.pk,
            usuario.nom_usua,
            usuario.apell_usua,
            usuario.email,
            usuario.rol,
        ])

    # Ajustar ancho de columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Preparar respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'
    wb.save(response)

    return response


def guardar_usuario(request):
    if request.method == 'POST':
        nombre = request.POST['nomUsua']
        apellido = request.POST['apellUsua']
        email = request.POST['emailUsua']
        password = request.POST['passwUsua']
        rol = request.POST['rol']
        activo = request.POST['activo']
        # Aquí deberías guardar el usuario, ejemplo:
        # Usuario.objects.create(...)
        # ...existing code...

@login_required
def exportar_excel(request):
    usuario = request.user
    fecha_desde = request.GET.get('desde')
    fecha_hasta = request.GET.get('hasta')

    envios = Envio.objects.filter(cod_domi=usuario.domiciliario)
    if fecha_desde:
        envios = envios.filter(fecha_entrega__date__gte=fecha_desde)
    if fecha_hasta:
        envios = envios.filter(fecha_entrega__date__lte=fecha_hasta)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Código", "Fecha Entrega", "Tarifa", "Estado"])

    for envio in envios:
        ws.append([envio.cod_envio, str(envio.fecha_entrega), envio.tarifa_envio, envio.estado])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="historial_envios.xlsx"'
    wb.save(response)
    return response

@login_required
def exportar_pdf(request):
    usuario = request.user
    fecha_desde = request.GET.get('desde')
    fecha_hasta = request.GET.get('hasta')
    envios = Envio.objects.filter(cod_domi=usuario.domiciliario).order_by('fecha_entrega')

    if fecha_desde not in [None, '', 'None']:
        envios = envios.filter(fecha_entrega__date__gte=fecha_desde)
    if fecha_hasta not in [None, '', 'None']:
        envios = envios.filter(fecha_entrega__date__lte=fecha_hasta)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="historial_envios.pdf"'
    p = canvas.Canvas(response)
    y = 800
    p.setFont("Helvetica-Bold", 14)
    p.drawString(100, y, "Historial de Envíos")
    y -= 30
    p.setFont("Helvetica", 10)
    for envio in envios:
        fecha_str = envio.fecha_entrega.strftime("%Y-%m-%d %H:%M") if envio.fecha_entrega else "Sin entregar"
        p.drawString(100, y, f"{envio.cod_envio} | {fecha_str} | {envio.tarifa_envio} | {envio.estado}")
        y -= 20
        if y < 50:
            p.showPage()
            y = 800
            p.setFont("Helvetica", 10)
    p.save()
    return response

class NoCacheMiddleware:
    def __init__(self, get_response):
        self.get_response = get_response
    def __call__(self, request):
        response = self.get_response(request)
        response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

# ----------- REPORTES PDF (GRÁFICO DE BARRAS) -----------

@login_required
def reporte_insumo(request):
    insumos = Insumo.objects.all()
    nombres = [i.nomb_insumo for i in insumos]
    cantidades = [i.cnt_insumo for i in insumos]
    plt.figure(figsize=(8, 4))
    plt.bar(nombres, cantidades, color='#fc6998')
    plt.title('Cantidad de Insumos')
    plt.xlabel('Insumo')
    plt.ylabel('Cantidad')
    plt.xticks(rotation=30, ha='right')
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    grafico_base64 = base64.b64encode(buf.read()).decode('utf-8')
    buf.close()
    plt.close()
    context = {'insumos': insumos, 'grafico_base64': grafico_base64}
    html_string = render_to_string('reportes/reporte_insumo.html', context)
    pdf_file = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_insumos.pdf"'
    return response

@login_required
def reporte_entradas(request):
    entradas = Entrada.objects.all()
    fechas = [e.fecha_hora_entrada.strftime('%Y-%m-%d') for e in entradas]
    cantidades = [e.cnt_entrada for e in entradas]
    plt.figure(figsize=(8, 4))
    plt.bar(fechas, cantidades, color='#fc6998')
    plt.title('Entradas de Insumos')
    plt.xlabel('Fecha')
    plt.ylabel('Cantidad')
    plt.xticks(rotation=30, ha='right')
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    grafico_base64 = base64.b64encode(buf.read()).decode('utf-8')
    buf.close()
    plt.close()
    context = {'entradas': entradas, 'grafico_base64': grafico_base64}
    html_string = render_to_string('reportes/reporte_entradas.html', context)
    pdf_file = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_entradas.pdf"'
    return response

@login_required
def reporte_salidas(request):
    salidas = Salida.objects.all()
    fechas = [s.fecha_hora_salida.strftime('%Y-%m-%d') for s in salidas]
    cantidades = [s.cantidad for s in salidas]
    plt.figure(figsize=(8, 4))
    plt.bar(fechas, cantidades, color='#fc6998')
    plt.title('Salidas de Insumos')
    plt.xlabel('Fecha')
    plt.ylabel('Cantidad')
    plt.xticks(rotation=30, ha='right')
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    grafico_base64 = base64.b64encode(buf.read()).decode('utf-8')
    buf.close()
    plt.close()
    context = {'salidas': salidas, 'grafico_base64': grafico_base64}
    html_string = render_to_string('reportes/reporte_salidas.html', context)
    pdf_file = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_salidas.pdf"'
    return response

@login_required
def reporte_ventas(request):
    ventas = Venta.objects.all()
    fechas = [v.fecha_hora.strftime('%Y-%m-%d') for v in ventas]
    totales = [float(v.total) for v in ventas]
    plt.figure(figsize=(8, 4))
    plt.bar(fechas, totales, color='#fc6998')
    plt.title('Ventas')
    plt.xlabel('Fecha')
    plt.ylabel('Total ($)')
    plt.xticks(rotation=30, ha='right')
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    grafico_base64 = base64.b64encode(buf.read()).decode('utf-8')
    buf.close()
    plt.close()
    context = {'ventas': ventas, 'grafico_base64': grafico_base64}
    html_string = render_to_string('reportes/reporte_ventas.html', context)
    pdf_file = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_ventas.pdf"'
    return response

@login_required
def reporte_produccion(request):
    producciones = Produccion.objects.all()
    estados = [p.get_estado_display() for p in producciones]
    from collections import Counter
    conteo = Counter(estados)
    plt.figure(figsize=(6, 4))
    plt.bar(conteo.keys(), conteo.values(), color='#fc6998')
    plt.title('Estados de Producción')
    plt.xlabel('Estado')
    plt.ylabel('Cantidad')
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    grafico_base64 = base64.b64encode(buf.read()).decode('utf-8')
    buf.close()
    plt.close()
    context = {'producciones': producciones, 'grafico_base64': grafico_base64}
    html_string = render_to_string('reportes/reporte_produccion.html', context)
    pdf_file = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_produccion.pdf"'
    return response

@login_required
def reporte_proveedores(request):
    proveedores = Proveedor.objects.all()
    nombres = [p.nom_proveedor for p in proveedores]
    entradas = [p.entradas.count() for p in proveedores]
    plt.figure(figsize=(8, 4))
    plt.bar(nombres, entradas, color='#fc6998')
    plt.title('Entradas por Proveedor')
    plt.xlabel('Proveedor')
    plt.ylabel('Entradas')
    plt.xticks(rotation=30, ha='right')
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    grafico_base64 = base64.b64encode(buf.read()).decode('utf-8')
    buf.close()
    plt.close()
    context = {'proveedores': proveedores, 'grafico_base64': grafico_base64}
    html_string = render_to_string('reportes/reporte_proveedores.html', context)
    pdf_file = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_proveedores.pdf"'
    return response

@login_required
def reporte_envios(request):
    envios = Envio.objects.all()
    estados = [e.get_estado_display() for e in envios]
    from collections import Counter
    conteo = Counter(estados)
    plt.figure(figsize=(6, 4))
    plt.bar(conteo.keys(), conteo.values(), color='#fc6998')
    plt.title('Estados de Envíos')
    plt.xlabel('Estado')
    plt.ylabel('Cantidad')
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    grafico_base64 = base64.b64encode(buf.read()).decode('utf-8')
    buf.close()
    plt.close()
    context = {'envios': envios, 'grafico_base64': grafico_base64}
    html_string = render_to_string('reportes/reporte_envios.html', context)
    pdf_file = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_envios.pdf"'
    return response

@login_required
def reporte_categorias(request):
    categorias = CategoriaInsumo.objects.all()
    nombres = [c.nom_categoria for c in categorias]
    insumos_count = [c.insumos.count() for c in categorias]
    plt.figure(figsize=(8, 4))
    plt.bar(nombres, insumos_count, color='#fc6998')
    plt.title('Cantidad de Insumos por Categoría')
    plt.xlabel('Categoría')
    plt.ylabel('Cantidad de Insumos')
    plt.xticks(rotation=30, ha='right')
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    grafico_base64 = base64.b64encode(buf.read()).decode('utf-8')
    buf.close()
    plt.close()
    context = {'categorias': categorias, 'grafico_base64': grafico_base64}
    html_string = render_to_string('reportes/reporte_categorias.html', context)
    pdf_file = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_categorias.pdf"'
    return response